from flask import Flask, request, render_template, jsonify, send_file
import os
import tempfile
import openpyxl
import pydicom
from dicompylercore import dicomparser, dvhcalc, dvh
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import sys
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'dcm'}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/faq')
def faq():
    return render_template('faq.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # Dateien aus dem Request holen
        rtdose_file = request.files.get('rtdose')
        rtstruct_file = request.files.get('rtstruct')
        is_example = request.form.get('is_example') == 'true'
        
        if is_example:
            # Example-Dateipfade
            example_path = os.path.join(app.root_path, 'exampleDCM')
            rtdose_path = os.path.join(example_path, 'RTDOSE_0.dcm')
            rtstruct_path = os.path.join(example_path, 'RTSTRUCT_0.dcm')
            
            if not os.path.exists(rtdose_path) or not os.path.exists(rtstruct_path):
                return jsonify({'error': 'Example-Dateien nicht gefunden'}), 404
                
            # DICOM-Dateien einlesen
            rtdose = pydicom.dcmread(rtdose_path, force=True)
            rtstruct = pydicom.dcmread(rtstruct_path, force=True)
        else:
            if not rtdose_file or not rtstruct_file:
                return jsonify({'error': 'Beide DICOM-Dateien sind erforderlich'}), 400

            # Temporäre Dateien speichern
            temp_rtdose = tempfile.NamedTemporaryFile(delete=False)
            temp_rtstruct = tempfile.NamedTemporaryFile(delete=False)
            
            try:
                rtdose_file.save(temp_rtdose.name)
                rtstruct_file.save(temp_rtstruct.name)
                
                # DICOM-Dateien einlesen
                rtdose = pydicom.dcmread(temp_rtdose.name, force=True)
                rtstruct = pydicom.dcmread(temp_rtstruct.name, force=True)
            finally:
                # Temporäre Dateien löschen
                try:
                    os.unlink(temp_rtdose.name)
                    os.unlink(temp_rtstruct.name)
                except Exception as e:
                    print(f"Warnung: Konnte temporäre Datei nicht löschen: {str(e)}")

        # Überprüfen der DICOM-Modalitäten
        if rtdose.Modality != 'RTDOSE':
            raise ValueError('Erste Datei ist keine RTDOSE-Datei')
        if rtstruct.Modality != 'RTSTRUCT':
            raise ValueError('Zweite Datei ist keine RTSTRUCT-Datei')
        
        # DVH berechnen
        rtdose_dcm = dicomparser.DicomParser(rtdose)
        rtstruct_dcm = dicomparser.DicomParser(rtstruct)
        structures = rtstruct_dcm.GetStructures()
        
        dvh_data = {}
        error_structures = {}
        
        # ROI-Mapping erstellen
        roi_mapping = {}
        for roi_contour in rtstruct.ROIContourSequence:
            if hasattr(roi_contour, 'ContourSequence'):
                roi_mapping[roi_contour.ReferencedROINumber] = len(roi_contour.ContourSequence)
        
        for structure_id in structures:
            structure = structures[structure_id]
            structure_name = structure['name']
            
            # Debug-Ausgabe
            print(f"\nPrüfe Struktur: {structure_name}")
            
            # Strukturen filtern (case-insensitive)
            name_lower = structure_name.lower().replace(' ', '')  # Entferne Leerzeichen für den Vergleich
            print(f"Name nach Konvertierung: {name_lower}")
            
            # Prüfe zuerst auf ausgeschlossene Präfixe
            should_skip = False
            if structure_name.startswith('_'):
                print(f"Skip wegen '_' Prefix")
                should_skip = True
            elif name_lower.startswith(('enc', 'dose', 'globe')):
                print(f"Skip wegen Prefix: {name_lower}")
                should_skip = True
            elif 'density' in name_lower:
                print(f"Skip wegen 'density'")
                should_skip = True
            elif 'brain-gtv' in name_lower:
                print(f"Skip wegen 'brain-gtv'")
                should_skip = True
            elif structure['type'].upper() in ['EXTERNAL', 'ERROR']:
                print(f"Skip wegen Typ: {structure['type']}")
                should_skip = True
                
            if should_skip:
                print(f"Überspringe Struktur: {structure_name}")
                continue
            
            try:
                print(f"\nVerarbeite Struktur: {structure_name}")
                
                # Konturinformationen aus dem Mapping
                if structure['id'] in roi_mapping:
                    print(f"  Anzahl Konturen: {roi_mapping[structure['id']]}")
                else:
                    print("  Keine Konturen gefunden")
                    continue
                        
                print(f"  Berechne DVH für Struktur {structure_name}...")
                
                # Berechne DVH
                limit = None  # Kein Limit für die Dosis
                
                # Mehr Interpolation für kleine Strukturen
                if structure['id'] in roi_mapping and roi_mapping[structure['id']] < 15:
                    print(f"  Struktur hat weniger als 15 Konturen ({roi_mapping[structure['id']]}), erhöhe Interpolation...")
                    dvh = dvhcalc.get_dvh(rtstruct, rtdose, structure_id, limit=limit, interpolation_segments_between_planes=2)
                else:
                    dvh = dvhcalc.get_dvh(rtstruct, rtdose, structure_id, limit=limit)
                
                if dvh is None:
                    error_structures[structure_name] = "DVH-Berechnung fehlgeschlagen"
                    continue
                        
                if dvh.volume <= 0:
                    error_structures[structure_name] = "Struktur hat kein Volumen"
                    continue
                    
                # Sicheres Konvertieren der DVH-Werte
                def safe_float(value):
                    try:
                        if hasattr(value, 'value'):  # Für DVHValue Objekte
                            return float(value.value)
                        return float(value)
                    except:
                        return 0.0
                    
                # Behalte cGy bei
                dvh_data[structure_name] = {
                    'doses': [float(x) for x in dvh.bincenters],  
                    'volumes': [float(x) for x in (dvh.counts / dvh.counts[0] * 100)],  
                    'volume': safe_float(dvh.volume),
                    'min_dose': safe_float(dvh.min),  
                    'max_dose': safe_float(dvh.max),  
                    'mean_dose': safe_float(dvh.mean),  
                    'D100': safe_float(dvh.D100),  
                    'D98': safe_float(dvh.D98),  
                    'D95': safe_float(dvh.D95),  
                    'D2cc': safe_float(dvh.D2cc)  
                }
                print(f"  DVH erfolgreich berechnet")
                print(f"  Volume: {dvh_data[structure_name]['volume']:.2f} cm³")
                print(f"  Min/Max Dose: {dvh_data[structure_name]['min_dose']:.2f}/{dvh_data[structure_name]['max_dose']:.2f} cGy")
                    
            except Exception as e:
                error_msg = str(e)
                print(f"  DVH-Berechnungsfehler: {error_msg}")
                error_structures[structure_name] = error_msg
                continue
        
        # Referenzen freigeben
        rtdose = None
        rtstruct = None
        rtdose_dcm = None
        rtstruct_dcm = None
        
        if not dvh_data:
            if error_structures:
                error_msg = "Fehler bei der DVH-Berechnung:\\n"
                for struct, err in error_structures.items():
                    error_msg += f"- {struct}: {err}\\n"
                return jsonify({'error': error_msg}), 400
            return jsonify({'error': 'Keine gültigen Strukturen gefunden.'}), 400
        
        # Rückgabe der erfolgreichen Daten und Fehlermeldungen
        return jsonify({
            'dvh_data': dvh_data,
            'errors': error_structures
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
    finally:
        # Warte kurz, um sicherzustellen, dass alle Referenzen freigegeben sind
        import time
        time.sleep(0.5)

@app.route('/load_example', methods=['POST'])
def load_example():
    try:
        example_path = os.path.join(app.root_path, 'exampleDCM')
        rtdose_path = os.path.join(example_path, 'RTDOSE_0.dcm')
        rtstruct_path = os.path.join(example_path, 'RTSTRUCT_0.dcm')
        
        if not os.path.exists(rtdose_path) or not os.path.exists(rtstruct_path):
            return jsonify({'error': 'Example-Dateien nicht gefunden'}), 404
            
        # DICOM-Dateien einlesen
        rtdose = pydicom.dcmread(rtdose_path, force=True)
        rtstruct = pydicom.dcmread(rtstruct_path, force=True)
        
        # DVH berechnen
        rtdose_dcm = dicomparser.DicomParser(rtdose)
        rtstruct_dcm = dicomparser.DicomParser(rtstruct)
        structures = rtstruct_dcm.GetStructures()
        
        dvh_data = {}
        error_structures = {}
        
        # ROI-Mapping erstellen
        roi_mapping = {}
        for roi_contour in rtstruct.ROIContourSequence:
            if hasattr(roi_contour, 'ContourSequence'):
                roi_mapping[roi_contour.ReferencedROINumber] = len(roi_contour.ContourSequence)
        
        for structure_id in structures:
            structure = structures[structure_id]
            structure_name = structure['name']
            
            # Debug-Ausgabe
            print(f"\nPrüfe Struktur: {structure_name}")
            
            # Strukturen filtern (case-insensitive)
            name_lower = structure_name.lower().replace(' ', '')  # Entferne Leerzeichen für den Vergleich
            print(f"Name nach Konvertierung: {name_lower}")
            
            # Prüfe zuerst auf ausgeschlossene Präfixe
            should_skip = False
            if structure_name.startswith('_'):
                print(f"Skip wegen '_' Prefix")
                should_skip = True
            elif name_lower.startswith(('enc', 'dose', 'globe')):
                print(f"Skip wegen Prefix: {name_lower}")
                should_skip = True
            elif 'density' in name_lower:
                print(f"Skip wegen 'density'")
                should_skip = True
            elif structure['type'].upper() in ['EXTERNAL', 'ERROR']:
                print(f"Skip wegen Typ: {structure['type']}")
                should_skip = True
                
            if should_skip:
                print(f"Überspringe Struktur: {structure_name}")
                continue
            
            try:
                print(f"\nVerarbeite Struktur: {structure_name}")
                
                # Konturinformationen aus dem Mapping
                if structure['id'] in roi_mapping:
                    print(f"  Anzahl Konturen: {roi_mapping[structure['id']]}")
                else:
                    print("  Keine Konturen gefunden")
                    continue
                    
                print(f"  Berechne DVH für Struktur {structure_name}...")
                
                # Berechne DVH
                limit = None  # Kein Limit für die Dosis
                
                # Mehr Interpolation für kleine Strukturen
                if structure['id'] in roi_mapping and roi_mapping[structure['id']] < 15:
                    print(f"  Struktur hat weniger als 15 Konturen ({roi_mapping[structure['id']]}), erhöhe Interpolation...")
                    dvh = dvhcalc.get_dvh(rtstruct, rtdose, structure_id, limit=limit, interpolation_segments_between_planes=2)
                else:
                    dvh = dvhcalc.get_dvh(rtstruct, rtdose, structure_id, limit=limit)
                
                if dvh is None:
                    error_structures[structure_name] = "DVH-Berechnung fehlgeschlagen"
                    continue
                    
                print("  DVH erfolgreich berechnet")
                
                # Daten für die Struktur speichern
                dvh_data[structure_name] = {
                    'doses': [float(x) for x in dvh.bincenters],
                    'volumes': [float(x) for x in dvh.counts],
                    'volume': float(dvh.volume),
                    'min_dose': float(dvh.min),
                    'max_dose': float(dvh.max),
                    'mean_dose': float(dvh.mean),
                    'D100': float(dvh.D100.value) if hasattr(dvh.D100, 'value') else float(dvh.D100),
                    'D98': float(dvh.D98.value) if hasattr(dvh.D98, 'value') else float(dvh.D98),
                    'D95': float(dvh.D95.value) if hasattr(dvh.D95, 'value') else float(dvh.D95),
                    'D2cc': float(dvh.D2cc.value) if hasattr(dvh, 'D2cc') and hasattr(dvh.D2cc, 'value') else 0
                }
                
            except Exception as e:
                print(f"Fehler bei Struktur {structure_name}: {str(e)}")
                error_structures[structure_name] = str(e)
        
        if not dvh_data:
            return jsonify({'error': 'Keine DVH-Daten gefunden'}), 400
            
        return jsonify({
            'dvh_data': dvh_data,
            'error_structures': error_structures
        })
        
    except Exception as e:
        print(f"Fehler beim Laden der Example-Daten: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export', methods=['POST'])
def export_data():
    try:
        data = request.get_json()
        dvh_data = data.get('dvh_data', {})
        
        if not dvh_data:
            return jsonify({'error': 'Keine Daten zum Exportieren'}), 400

        # Excel-Datei erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "DVH Metrics"
        
        # Header
        headers = ['Structure', 'Volume (cc)', 'Min Dose (Gy)', 'Max Dose (Gy)', 'Mean Dose (Gy)',
                  'D100 (Gy)', 'D98 (Gy)', 'D95 (Gy)', 'D2cc (Gy)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Daten einfügen
        row = 2
        for structure, metrics in dvh_data.items():
            ws.cell(row=row, column=1, value=structure)
            ws.cell(row=row, column=2, value=round(metrics['volume'], 3))
            ws.cell(row=row, column=3, value=round(metrics['min_dose'], 2))
            ws.cell(row=row, column=4, value=round(metrics['max_dose'], 2))
            ws.cell(row=row, column=5, value=round(metrics['mean_dose'], 2))
            ws.cell(row=row, column=6, value=round(metrics['D100'], 2))
            ws.cell(row=row, column=7, value=round(metrics['D98'], 2))
            ws.cell(row=row, column=8, value=round(metrics['D95'], 2))
            ws.cell(row=row, column=9, value=round(metrics['D2cc'], 2))
            row += 1
        
        # Zweites Worksheet für DVH-Daten
        ws_dvh = wb.create_sheet("DVH Data")
        ws_dvh.cell(row=1, column=1, value="Dose (Gy)")
        
        # Erste Spalte: Dosiswerte
        first_structure = list(dvh_data.keys())[0]
        doses = dvh_data[first_structure]['doses']
        for i, dose in enumerate(doses, 2):
            ws_dvh.cell(row=i, column=1, value=round(dose, 4))  # Erhöhe Genauigkeit auf 4 Dezimalstellen
        
        # Weitere Spalten: Volumen pro Struktur
        col = 2
        for structure, data in dvh_data.items():
            ws_dvh.cell(row=1, column=col, value=f"{structure} Volume (%)")
            volumes = data['volumes']
            for i, vol in enumerate(volumes, 2):
                ws_dvh.cell(row=i, column=col, value=round(vol, 3))  # Volumen auf 3 Dezimalstellen
            col += 1
        
        # Datei speichern und zurückgeben
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Zeitstempel für Dateinamen
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'dvh_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print("Export error:", str(e))
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("Starting Flask application...")
    sys.stdout.flush()
    app.run(debug=True)
